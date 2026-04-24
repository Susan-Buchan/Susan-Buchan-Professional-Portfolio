"""
Data Cleaning Script for North Carolina Public Library Statistical Reports
=========================================================================

This script cleans Excel workbooks containing library data by:
1. Removing blank columns
2. Removing header/title rows
3. Unmerging cells (handled by pandas when reading)
4. Creating a single header row with clean column names

Usage:
    python clean_library_data.py input_file.xlsx output_file.xlsx
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import re


def detect_header_row(df, max_rows_to_check=10):
    """
    Detect the actual header row by looking for rows with mostly non-null string values
    that look like column headers (not numeric data).
    """
    for idx in range(min(max_rows_to_check, len(df))):
        row = df.iloc[idx]
        non_null_count = row.notna().sum()
        
        # Check if this row has enough non-null values
        if non_null_count >= len(df.columns) * 0.5:
            # Check if values look like headers (mostly strings, not numbers)
            string_count = sum(1 for val in row if isinstance(val, str) and not val.replace(',', '').replace('.', '').replace('$', '').replace('%', '').lstrip('-').isdigit())
            if string_count >= non_null_count * 0.6:
                return idx
    return 0


def detect_data_start_row(df, header_row):
    """
    Detect where the actual data starts (after any sub-headers or blank rows).
    """
    for idx in range(header_row + 1, min(header_row + 5, len(df))):
        row = df.iloc[idx]
        non_null_count = row.notna().sum()
        if non_null_count >= len(df.columns) * 0.3:
            return idx
    return header_row + 1


def remove_blank_columns(df, threshold=0.9):
    """
    Remove columns that are mostly empty (>threshold proportion null).
    """
    null_ratios = df.isnull().sum() / len(df)
    cols_to_keep = null_ratios[null_ratios < threshold].index
    return df[cols_to_keep]


def remove_blank_rows(df, threshold=0.9):
    """
    Remove rows that are mostly empty (>threshold proportion null).
    """
    null_ratios = df.isnull().sum(axis=1) / len(df.columns)
    return df[null_ratios < threshold]


def clean_header_names(headers):
    """
    Clean header names by removing extra whitespace and standardizing format.
    """
    cleaned = []
    for h in headers:
        if pd.isna(h):
            cleaned.append('')
        else:
            # Convert to string, strip whitespace, collapse multiple spaces
            h_clean = re.sub(r'\s+', ' ', str(h).strip())
            cleaned.append(h_clean)
    return cleaned


def merge_multi_row_headers(df, header_rows):
    """
    Merge multiple header rows into a single header row.
    
    Args:
        df: DataFrame with raw data
        header_rows: List of row indices that contain header information
    
    Returns:
        List of merged header names
    """
    if len(header_rows) == 0:
        return list(df.columns)
    
    merged_headers = []
    for col_idx in range(len(df.columns)):
        parts = []
        for row_idx in header_rows:
            if row_idx < len(df):
                val = df.iloc[row_idx, col_idx]
                if pd.notna(val) and str(val).strip():
                    parts.append(str(val).strip())
        
        # Join non-empty parts with space
        header = ' '.join(parts) if parts else f'Column_{col_idx + 1}'
        merged_headers.append(header)
    
    return merged_headers


def clean_sheet(df, sheet_name=''):
    """
    Clean a single sheet from the library data workbook.
    
    Args:
        df: Raw DataFrame from Excel
        sheet_name: Name of the sheet (for logging)
    
    Returns:
        Cleaned DataFrame
    """
    print(f"\nCleaning sheet: {sheet_name}")
    print(f"  Original shape: {df.shape}")
    
    # Step 1: Remove completely empty columns
    df = remove_blank_columns(df, threshold=0.95)
    print(f"  After removing blank columns: {df.shape}")
    
    # Step 2: Detect title/header rows to skip
    # Look for rows that contain title keywords
    title_keywords = ['table', 'statistical report', 'library', 'profile', 
                      'north carolina', 'fiscal year', 'fy 20', 'fy20']
    
    rows_to_skip = []
    for idx in range(min(10, len(df))):
        row_str = ' '.join([str(v).lower() for v in df.iloc[idx] if pd.notna(v)])
        if any(keyword in row_str for keyword in title_keywords):
            rows_to_skip.append(idx)
        elif df.iloc[idx].isna().sum() >= len(df.columns) * 0.8:
            rows_to_skip.append(idx)
    
    # Step 3: Find the actual header row (first row with column names)
    header_row = None
    for idx in range(min(10, len(df))):
        if idx in rows_to_skip:
            continue
        row = df.iloc[idx]
        non_null = row.notna().sum()
        if non_null >= len(df.columns) * 0.4:
            # Check if this looks like a header (contains text like 'Library', 'FSCS', etc.)
            row_text = ' '.join([str(v) for v in row if pd.notna(v)])
            if any(word in row_text for word in ['Library', 'FSCS', 'Type', 'Name', 'County']):
                header_row = idx
                break
    
    if header_row is None:
        header_row = 0
    
    print(f"  Detected header row: {header_row}")
    
    # Step 4: Check for multi-row headers
    header_rows = []
    for idx in range(max(0, header_row - 2), header_row + 1):
        if idx not in rows_to_skip:
            row = df.iloc[idx]
            non_null = row.notna().sum()
            if non_null >= 2:  # At least some values
                header_rows.append(idx)
    
    # Step 5: Create merged headers
    if len(header_rows) > 1:
        new_headers = merge_multi_row_headers(df, header_rows)
    else:
        new_headers = [str(v) if pd.notna(v) else f'Column_{i}' 
                       for i, v in enumerate(df.iloc[header_row])]
    
    new_headers = clean_header_names(new_headers)
    
    # Step 6: Get data rows (after headers)
    data_start = max(header_rows) + 1 if header_rows else header_row + 1
    df_clean = df.iloc[data_start:].copy()
    df_clean.columns = new_headers
    df_clean = df_clean.reset_index(drop=True)
    
    # Step 7: Remove any remaining blank rows
    df_clean = remove_blank_rows(df_clean, threshold=0.9)
    
    # Step 8: Remove any unnamed columns or columns with no header
    cols_to_keep = [col for col in df_clean.columns if col and not col.startswith('Column_')]
    if len(cols_to_keep) < len(df_clean.columns):
        # Keep original columns if too many would be removed
        if len(cols_to_keep) >= len(df_clean.columns) * 0.5:
            df_clean = df_clean[cols_to_keep]
    
    print(f"  Final shape: {df_clean.shape}")
    print(f"  Headers: {list(df_clean.columns[:5])}...")
    
    return df_clean


def clean_workbook(input_file, output_file=None):
    """
    Clean all sheets in a library data workbook.
    
    Args:
        input_file: Path to input Excel file
        output_file: Path for output file (optional, defaults to input_file with '_cleaned' suffix)
    
    Returns:
        Dictionary of cleaned DataFrames
    """
    if output_file is None:
        output_file = input_file.replace('.xlsx', '_cleaned.xlsx')
    
    print(f"Reading workbook: {input_file}")
    
    # Read all sheets
    xl = pd.ExcelFile(input_file)
    sheet_names = xl.sheet_names
    print(f"Found {len(sheet_names)} sheets: {sheet_names}")
    
    cleaned_sheets = {}
    
    for sheet_name in sheet_names:
        # Read sheet without headers (we'll detect them)
        df = pd.read_excel(input_file, sheet_name=sheet_name, header=None)
        
        # Clean the sheet
        df_clean = clean_sheet(df, sheet_name)
        cleaned_sheets[sheet_name] = df_clean
    
    # Write cleaned data to new workbook
    print(f"\nWriting cleaned data to: {output_file}")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name, df in cleaned_sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Apply formatting to headers
    format_headers(output_file)
    
    print("\nCleaning complete!")
    return cleaned_sheets


def format_headers(file_path):
    """
    Apply professional formatting to header rows in the workbook.
    """
    wb = load_workbook(file_path)
    
    # Define header style
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Format header row (row 1)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-fit column widths (approximate)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(file_path)
    print(f"Applied formatting to headers")


# Example usage and specific sheet configurations
SHEET_CONFIGS = {
    'Table 1': {
        'expected_headers': ['Library Type', 'NC Dept. of Commerce Tier Designation (2020)', 
                            'Legal Service Population Area', 'Central', 'Branches', 
                            'Bookmobiles', 'Mobile Units', 'Kiosks', 'Annual Hours', 
                            'Library Square Feet per Capita'],
        'skip_rows': 6
    },
    'Table 2 - Staff': {
        'expected_headers': ['FSCS Key', 'Library Name', 'Library Type', 'FTE ALA/MLS', 
                            'FTE MLS Not ALA', 'Total FTE MLS', 'Other Paid Staff', 
                            'Total FTE Staff', 'FTE Per 25000 Population', 
                            '% of Staff with ALA/MLS', 'Volunteer Hours'],
        'skip_rows': 5
    },
    # Add more sheet configurations as needed
}


def main():
    """
    Main entry point for the cleaning script.
    """
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python clean_library_data.py <input_file.xlsx> [output_file.xlsx]")
        print("\nExample:")
        print("  python clean_library_data.py nc_library_stats.xlsx nc_library_stats_clean.xlsx")
        return
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    clean_workbook(input_file, output_file)


if __name__ == '__main__':
    main()